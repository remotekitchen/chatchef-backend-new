# app/services/invoice_vr_excel.py
import io, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

# Aliases coming from Lark (case-insensitive); we normalize to the keys below
ALIAS_MAP = {
    "Revenue GST": ["Revenue GST", "revenue gst", "GST (revenue)", "gst revenue", "gst"],
    "Revenue PST": ["Revenue PST", "revenue pst", "PST (revenue)", "pst revenue", "pst"],
    "Markup":      ["Markup", "markup", "Markup (10%)", "markup (10%)"],
}

# Final preferred order in the Excel (VR name + Platform remain at the end)
PREFERRED_ORDER = [
    "Date", "Description",
    "Markup",                      # <— normalized header
    "Discount",
    "Revenue GST", "Revenue PST",  # <— new columns from Lark
    "Platform fees (Commission + Tax)",
    "Selling price",
    "Payout from platforms", "Unit Price", "Sales Tax", "Total",
    "VR name", "Platform",
]

def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename any alias found in ALIAS_MAP to its canonical header.
    We only rename the *first* alias match to avoid duplicates.
    """
    cols_lower = {c.lower(): c for c in df.columns}
    renames = {}
    for canonical, aliases in ALIAS_MAP.items():
        for a in aliases:
            src = cols_lower.get(a.lower())
            if src and canonical not in df.columns:
                renames[src] = canonical
                break
    if renames:
        df = df.rename(columns=renames)
    return df

def _autosize(ws):
    for col_cells in ws.columns:
        letter = next((c.column_letter for c in col_cells if not isinstance(c, MergedCell)), None)
        if not letter:
            continue
        width = max((len(str(c.value)) for c in col_cells if c.value), default=0) + 2
        ws.column_dimensions[letter].width = width

def build_vr_excel(df: pd.DataFrame) -> bytes:
    # Normalize incoming Lark columns to our canonical headers
    df = _normalize_columns(df)

    # Reorder columns per layout; keep any extra columns before VR name/Platform block
    # (but after our main block), without duplicating VR name/Platform.
    main_cols = [c for c in PREFERRED_ORDER if c in df.columns and c not in ("VR name", "Platform")]
    trailing_cols = []
    if "VR name" in df.columns:
        trailing_cols.append("VR name")
    if "Platform" in df.columns:
        trailing_cols.append("Platform")

    # "Others" are those not already chosen
    chosen = set(main_cols + trailing_cols)
    others = [c for c in df.columns if c not in chosen]

    final_cols = main_cols + others + trailing_cols
    df = df[final_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    center = Alignment(horizontal="center", vertical="center")

    # Title row & date -> adapt to actual number of columns
    last_col = max(1, len(df.columns))
    last_letter = get_column_letter(last_col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = "INVOICE"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Date (top-right)
    ws[f"{last_letter}2"] = pd.Timestamp.now().strftime("%B %d, %Y")
    ws[f"{last_letter}2"].alignment = center

    # Pay to / From blocks
    ws["A4"] = "Pay to"
    ws["A5"] = "Nice Meeting You"
    ws["A6"] = "535 Clarke Rd, Coquitlam, BC V3J 3X4"

    ws["H4"] = "From"
    ws["H5"] = "Thunder Digital Kitchen"
    ws["H6"] = "200 - 13571 COMMERCE PKY, RICHMOND BC V6V 2R2, CANADA"

    ws.append([]); ws.append([])

    # Charges heading
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=last_col)
    ws["A9"] = "Charges"; ws["A9"].font = Font(bold=True)

    # Headers
    ws.append(list(df.columns))
    for i in range(1, len(df.columns) + 1):
        ws.cell(row=ws.max_row, column=i).font = Font(bold=True)

    # Rows
    grand_total = 0.0
    for _, r in df.iterrows():
        vals = []
        for col in df.columns:
            v = r.get(col)
            if col.lower() == "date":
                try:
                    v = pd.to_datetime(v).strftime("%Y-%m-%d")
                except:
                    pass
            vals.append(v)
            if col.lower() == "total":
                try:
                    grand_total += float(v or 0)
                except:
                    pass
        ws.append(vals)

    ws.append([]); ws.append([])

    # Totals line (based on "Total" column if present)
    if any(c.lower() == "total" for c in df.columns):
        pad = max(0, len(df.columns) - 2)
        ws.append([""] * pad + ["Amount to be Paid", f"{grand_total:.2f}"])
        ws.cell(row=ws.max_row, column=len(df.columns)).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=len(df.columns)-1).font = Font(bold=True)

    _autosize(ws)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()
