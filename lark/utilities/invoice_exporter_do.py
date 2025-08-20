import io
import os
import re
import datetime as dt
from typing import List, Dict, Tuple

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
try:
    from openpyxl.drawing.image import Image as XLImage
except Exception:
    XLImage = None

# Optional PDF helper (kept as-is)
try:
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

LOGO_PATH_TECHCHEF = os.getenv("LOGO_PATH_TECHCHEF", "").strip()
UTC8 = dt.timedelta(hours=8)  # Lark 'Date' normalization

# -------------------- small utils --------------------
def _slug(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^A-Za-z0-9\-_]+", "_", s)
    s = re.sub(r"_{2,}", "_", s).strip("_")
    return s or "unknown"

def _norm_text(v):
    if isinstance(v, dict) and "text" in v: return v["text"]
    if isinstance(v, list) and v and isinstance(v[0], dict) and "text" in v[0]: return v[0]["text"]
    if isinstance(v, (int, float)): return str(v)
    if isinstance(v, str): return v.strip()
    return v

def _num(v, default=0.0):
    try:
        if v is None or v == "":
            return float(default)
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(",", ""))
        except Exception:
            return float(default)

def _lark_ms_to_iso(ms) -> str:
    """Add +8h and return YYYY-MM-DD (fixes 'one day earlier')."""
    try:
        ms = int(_norm_text(ms))
        dtu = dt.datetime.utcfromtimestamp(ms / 1000.0)
        return (dtu + UTC8).date().isoformat()
    except Exception:
        return ""

def _safe_get(row: Dict, *names, default=None):
    for n in names:
        if n in row and row[n] is not None:
            return row[n]
    return default

def _excel_header_row():
    return [
        "Order Date", "Order ID", "Item Price", "Discount", "Payment Type", "Order Mode", "tax",
        "Selling price (inclusive of tax)", "Original Delivery Fees", "Customer absorb on delivery fees",
        "Delivery fees expense", "Stripe Fees", "service fees to restaurant", "service fee to techchef",
        "tips for restaurant", "bag fees", "utensil fees", "Refund Amount", "Sub-Total Payment",
    ]

# -------------------- Excel building --------------------
def _append_top_block(sh, restaurant_name: str):
    sh.append([]); sh.append([]); sh.append([]); sh.append([]); sh.append([])
    sh.append(["", "", "", "", "", "", "", "", "", ""])
    sh.append(["", "", "", "", "", "", "", "", "Date", f"{dt.datetime.today().date()}"])
    sh.append([]); sh.append([])
    sh.append(["Pay to"]); sh.append([f"{restaurant_name}"])
    sh.append([]); sh.append([])
    sh.append(["From"])
    sh.append(["Thunder Digital Kitchen Ltd"])
    sh.append(["200 - 13571 COMMERCE PKY, RICHMOND BC V6V 2R2, CANADA"])
    sh.append([])
    sh.append(["Charges"])
    sh.append(_excel_header_row())

def _style_and_logo(sh, payment_header_row: int):
    light_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)

    sh.column_dimensions['A'].width = 17
    sh.column_dimensions['B'].width = 37
    sh.column_dimensions['G'].width = 13
    sh.column_dimensions['I'].width = 17
    sh.column_dimensions['J'].width = 17

    rows_to_format = [10, 14, 19, payment_header_row]
    for row_idx in rows_to_format:
        try:
            for cell in sh[row_idx]:
                cell.fill = light_gray_fill
                cell.font = bold_font
        except Exception:
            pass

    try:
        sh.merge_cells('G1:K2')
        sh.merge_cells('A3:B9')
        sh.merge_cells(f'C{payment_header_row}:D{payment_header_row}')
        sh.merge_cells(f'C{payment_header_row+1}:D{payment_header_row+1}')
        sh.merge_cells(f'E{payment_header_row}:F{payment_header_row}')
        sh.merge_cells(f'E{payment_header_row+1}:F{payment_header_row+1}')
        sh.merge_cells(f'H{payment_header_row}:I{payment_header_row}')
        sh.merge_cells(f'H{payment_header_row+1}:I{payment_header_row+1}')
        sh.merge_cells(f'A{payment_header_row + 5}:B{payment_header_row + 5}')
    except Exception:
        pass

    sh['G1'] = "Order Summary"
    sh['G1'].font = Font(bold=True, size=20)
    sh['G1'].alignment = Alignment(horizontal="center", vertical="center")

    try:
        sh[f'A{payment_header_row - 1}'].font = Font(bold=True)
        sh[f'A{payment_header_row + 5}'].font = Font(bold=True, color="66D1EE", size=44)
    except Exception:
        pass

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sh.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

    if LOGO_PATH_TECHCHEF and os.path.exists(LOGO_PATH_TECHCHEF) and XLImage:
        try:
            img = XLImage(LOGO_PATH_TECHCHEF)
            img.width, img.height = 100, 100
            sh.add_image(img, 'A3')
        except Exception:
            pass

def _excel_bytes_do(rows: List[Dict], restaurant_name: str) -> bytes:
    wb = Workbook()
    sh = wb.active
    _append_top_block(sh, restaurant_name)

    cash_total = 0.0
    stripe_total = 0.0
    total = 0.0

    for r in rows:
        pay_type     = str(_norm_text(_safe_get(r, "Payment Type", "payment_type", ""))).upper()
        total_price  = _num(_safe_get(r, "Selling price (inclusive of tax)", "selling_price"))
        orig_deliv   = _num(_safe_get(r, "Original Delivery Fees", "Original_Delivery_Fees"))
        stripe_fees  = _num(_safe_get(r, "Stripe Fees", "stripe_fees"))
        svc_to_tech  = _num(_safe_get(r, "service fee to techchef", "Service fee to chatchefs", "Service fee to techchef"))
        refund_amt   = _num(_safe_get(r, "Refund Amount", "refund_amount"))

        sub_total = _safe_get(r, "Sub-Total Payment", "sub_total")
        if sub_total is None:
            sub_total = total_price - orig_deliv - stripe_fees - svc_to_tech - refund_amt
        sub_total = round(_num(sub_total), 2)

        sh.append([
            _lark_ms_to_iso(_safe_get(r, "Order Date")),
            str(_norm_text(_safe_get(r, "Order ID") or "")),
            f"{_num(_safe_get(r, 'Item Price')):.2f}",
            f"{_num(_safe_get(r, 'Discount')):.2f}",
            str(_norm_text(_safe_get(r, "Payment Type", "payment_type", ""))),
            str(_norm_text(_safe_get(r, "Order Mode") or "")),
            f"{_num(_safe_get(r, 'Tax', 'tax')):.2f}",
            f"{_num(_safe_get(r, 'Selling price (inclusive of tax)', 'selling_price')):.2f}",
            f"{orig_deliv:.2f}",
            f"{_num(_safe_get(r, 'Customer absorb on delivery fees', 'Customer_absorb_on_delivery_fees')):.2f}",
            f"{_num(_safe_get(r, 'Delivery fees expense', 'Delivery_fees_expense')):.2f}",
            f"{stripe_fees:.2f}",
            f"{_num(_safe_get(r, 'Service fees to restaurant', 'Service fees to Restaurant', 'service_fees_to_restaurant')):.2f}",
            f"{svc_to_tech:.2f}",
            f"{_num(_safe_get(r, 'Tips for restaurant', 'tips_for_restaurant')):.2f}",
            f"{_num(_safe_get(r, 'Bag fees', 'bag_fees')):.2f}",
            f"{_num(_safe_get(r, 'Utensil fees', 'utensil_fees')):.2f}",
            f"{refund_amt:.2f}",
            f"{sub_total:.2f}",
        ])

        total += sub_total
        if pay_type == "STRIPE":
            stripe_total += sub_total
        elif pay_type == "CASH":
            cash_total += sub_total

    sh.append([]); sh.append([])
    sh.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", f"{total: .2f}"])
    sh.append([])
    sh.append(["Payment"])
    sh.append([
        "", "Comment", "", "", "Sales through Pay-in-Person", "", "",
        "Sales Through Stripe", "", "Subtotal For Total Sales", "", "", "", "", "", "", "", "Amount"
    ])
    sh.append([
        "", "Stripe Payments Deposited", "", "",
        f"{cash_total:.2f}", "", "",
        f"{stripe_total:.2f}", "",
        f"{total:.2f}", "", "", "", "",
        f"{0.00:.2f}", "", "",
        f"{stripe_total + 0.00:.2f}"
    ])
    sh.append([])
    sh.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "*Rounding Difference"])
    sh.append([])
    sh.append(["techchef"])
    sh.append([])
    sh.append([])

    payment_header_row = sh.max_row - 6
    _style_and_logo(sh, payment_header_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# -------------------- PDF building --------------------
def _pdf_bytes_do(rows: List[Dict], restaurant_name: str, start_ms: int, end_ms: int) -> bytes:
    start_iso = _lark_ms_to_iso(start_ms)
    end_iso   = _lark_ms_to_iso(end_ms)

    cols = _excel_header_row()
    data = [cols]
    cash_total = stripe_total = total = 0.0

    for r in rows:
        pay_type = str(_norm_text(_safe_get(r, "Payment Type", "payment_type", ""))).upper()
        sub_total = _num(_safe_get(r, "Sub-Total Payment", "sub_total", default=0.0))
        total += sub_total
        if pay_type == "STRIPE":
            stripe_total += sub_total
        elif pay_type == "CASH":
            cash_total += sub_total

        data.append([
            _lark_ms_to_iso(_safe_get(r, "Order Date")),
            str(_norm_text(_safe_get(r, "Order ID") or "")),
            f"{_num(_safe_get(r, 'Item Price')):.2f}",
            f"{_num(_safe_get(r, 'Discount')):.2f}",
            str(_norm_text(_safe_get(r, "Payment Type", "payment_type", ""))),
            str(_norm_text(_safe_get(r, "Order Mode") or "")),
            f"{_num(_safe_get(r, 'Tax', 'tax')):.2f}",
            f"{_num(_safe_get(r, 'Selling price (inclusive of tax)', 'selling_price')):.2f}",
            f"{_num(_safe_get(r, 'Original Delivery Fees', 'Original_Delivery_Fees')):.2f}",
            f"{_num(_safe_get(r, 'Customer absorb on delivery fees', 'Customer_absorb_on_delivery_fees')):.2f}",
            f"{_num(_safe_get(r, 'Delivery fees expense', 'Delivery_fees_expense')):.2f}",
            f"{_num(_safe_get(r, 'Stripe Fees', 'stripe_fees')):.2f}",
            f"{_num(_safe_get(r, 'Service fees to restaurant', 'Service fees to Restaurant', 'service_fees_to_restaurant')):.2f}",
            f"{_num(_safe_get(r, 'service fee to techchef', 'Service fee to chatchefs', 'Service fee to techchef')):.2f}",
            f"{_num(_safe_get(r, 'Tips for restaurant', 'tips_for_restaurant')):.2f}",
            f"{_num(_safe_get(r, 'Bag fees', 'bag_fees')):.2f}",
            f"{_num(_safe_get(r, 'Utensil fees', 'utensil_fees')):.2f}",
            f"{_num(_safe_get(r, 'Refund Amount', 'refund_amount')):.2f}",
            f"{sub_total:.2f}",
        ])

    if REPORTLAB_OK:
        buf = io.BytesIO()
        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        story = [Paragraph(f"{restaurant_name} — {start_iso} to {end_iso}", styles["Heading2"]),
                 Paragraph("Charges", styles["Heading3"]), Spacer(1, 6)]
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(t)
        doc.build(story)
        return buf.getvalue()

    # Tiny fallback PDF (unchanged)
    buf = io.BytesIO()
    def w(b): buf.write(b if isinstance(b, (bytes, bytearray)) else b.encode("latin-1"))
    title = f"{restaurant_name} — {start_iso} to {end_iso}\n\n"
    flat = [", ".join(cols)]
    for row in data[1:]:
        flat.append(", ".join(row))
    text = title + "\n".join(flat)
    content = ("BT /F1 10 Tf 40 550 Td 14 TL (" + text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)").replace("\r", "").replace("\n", ") T* (") + ") Tj ET")
    w(b"%PDF-1.4\n"); offs=[]
    def obj(num, body):
        offs.append(buf.tell()); w(f"{num} 0 obj\n"); w(body); w("\nendobj\n")
    obj(1, "<< /Type /Catalog /Pages 2 0 R >>")
    obj(2, "<< /Type /Pages /Kids [3 0 R] /Count 1 >>")
    obj(3, "<< /Type /Page /Parent 2 0 R /MediaBox [0 0 842 595] /Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>")
    stream = f"<< /Length {len(content)} >>\nstream\n{content}\nendstream"
    obj(4, stream)
    obj(5, "<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>")
    xref_start = buf.tell()
    w("xref\n0 6\n0000000000 65535 f \n")
    for off in offs:
        w(f"{off:010} 00000 n \n")
    w(f"trailer << /Size 6 /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF")
    return buf.getvalue()

# -------------------- PUBLIC API --------------------
def generate_invoice_files_do(
    rows: List[Dict],
    *,
    restaurant_name: str,
    start_ms: int,
    end_ms: int,
) -> Tuple[str, str]:
    """
    Builds DO-style Excel & PDF and SAVES them to your S3-compatible storage (DO Spaces).
    Returns (excel_url, pdf_url).
    """
    start_iso = _lark_ms_to_iso(start_ms)   # use nominal UTC+8 dates for names
    end_iso   = _lark_ms_to_iso(end_ms)
    rslug = _slug(restaurant_name)
    base_prefix = f"invoices/DO/{rslug}/{start_iso}__{end_iso}"

    x_name = f"{base_prefix}/DO_{rslug}_{start_iso}__{end_iso}.xlsx"
    p_name = f"{base_prefix}/DO_{rslug}_{start_iso}__{end_iso}.pdf"

    excel_bytes = _excel_bytes_do(rows, restaurant_name)
    pdf_bytes   = _pdf_bytes_do(rows, restaurant_name, start_ms, end_ms)

    default_storage.save(x_name, ContentFile(excel_bytes))
    default_storage.save(p_name, ContentFile(pdf_bytes))

    excel_url = default_storage.url(x_name)
    pdf_url   = default_storage.url(p_name)
    return excel_url, pdf_url
