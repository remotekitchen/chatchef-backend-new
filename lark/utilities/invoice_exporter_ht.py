# billing/utilities/lark/invoice_exporter_ht.py
import io
import os
import re
import datetime as dt
from typing import List, Dict, Tuple

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

LOGO_PATH_HUNGRY = os.getenv("LOGO_PATH_HUNGRY", "").strip()

# Use business/local timezone for HT dates: default Asia/Dhaka = UTC+6.
# Set HT_TZ_HOURS=6 (or your actual offset) in the environment if needed.
HT_TZ_HOURS = int(os.getenv("HT_TZ_HOURS", "6"))

def _slug(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^A-Za-z0-9\-_]+", "_", s)
    s = re.sub(r"_{2,}", "_", s).strip("_")
    return s or "unknown"

def _norm_text(v):
    if isinstance(v, dict) and "text" in v: return v["text"]
    if isinstance(v, list) and v and isinstance(v[0], dict) and "text" in v[0]: return v[0]["text"]
    if isinstance(v, (int, float)): return str(v)
    if isinstance(v, str): return v.strip()
    return v

def _num(v, default=0.0):
    try:
        if v in (None, ""): return float(default)
        return float(v)
    except Exception:
        try: return float(str(v).replace(",", ""))
        except Exception: return float(default)

def _parse_any_date_to_ms_utc(d) -> int | None:
    """Accept ms/s epoch or common date strings and return UTC ms."""
    if d is None:
        return None
    if isinstance(d, (int, float)):
        v = int(d)
        return v * 1000 if v < 10_000_000_000 else v
    s = str(d).strip()
    if s.isdigit():
        v = int(s)
        return v * 1000 if v < 10_000_000_000 else v
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            base = dt.datetime.strptime(s[:10], fmt).replace(tzinfo=dt.timezone.utc)
            return int(base.timestamp() * 1000)
        except Exception:
            pass
    return None

UTC8 = dt.timedelta(hours=8)

def _lark_ms_to_nominal_date(ms: int | str) -> dt.date:
    ms = int(str(ms))
    dtu = dt.datetime.utcfromtimestamp(ms / 1000.0)  # ms -> UTC
    return (dtu + UTC8).date()                      # shift to UTC+8, take date

def _nominal_label(ms: int | str) -> str:
    return _lark_ms_to_nominal_date(ms).isoformat()

def _ms_to_local_date_iso(ms) -> str:
    """
    Convert Lark ms timestamp to *local* business date for HT (prevents day-back shift).
    Local offset configurable via HT_TZ_HOURS (default +6 for Dhaka).
    """
    try:
        ms = _parse_any_date_to_ms_utc(ms)
        if ms is None:
            return ""
        dt_utc = dt.datetime.utcfromtimestamp(ms / 1000.0)
        local  = dt_utc + dt.timedelta(hours=HT_TZ_HOURS)
        return local.date().isoformat()
    except Exception:
        return ""

HDR = [
    'Order Date','Order Id','Actual Item Price','Item Price','Discount',
    'Special Discount(HT)','BOGO Item inflation percentage','BOGO Discount','BOGO Loss',
    'Voucher discount','Voucher code','Discount bear by Restaurant','Discount bear by Hungrytiger',
    'BOGO Discount bear by Restaurant','BOGO Discount bear by Hungrytiger',
    'Payment Type','Order Mode','Tax','Selling price (inclusive of tax)',
    'Customer absorb on delivery fees','Delivery fees expense','Commission Percentage',
    'Commission Amount','Service fees to Restaurant','Service fee to Hungrytiger',
    'Tips for restaurant','Bag fees','Container fees','Amount to Restaurant','HT Profit','Restaurant Name'
]

def _append_top_block(sh, restaurant_name: str):
    sh.append([]); sh.append([]); sh.append([]); sh.append([]); sh.append([])
    sh.append([""]*10)
    sh.append([""]*9 + ["Date","", f"{dt.datetime.today().date()}","",""])
    sh.append([]); sh.append([])
    sh.append(["Pay to"]); sh.append([restaurant_name]); sh.append([]); sh.append([])
    sh.append(["From"]); sh.append(["Thunder Digital Kitchen Ltd"])
    sh.append(["200 - 13571 COMMERCE PKY, RICHMOND BC V6V 2R2, CANADA"])
    sh.append([]); sh.append(["Charges"]); sh.append(HDR)

def _style_sheet(sh):
    light_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold = Font(bold=True)
    sh.column_dimensions['A'].width = 17
    sh.column_dimensions['B'].width = 37
    sh.column_dimensions['G'].width = 13
    sh.column_dimensions['I'].width = 17
    sh.column_dimensions['J'].width = 17
    for row_idx in (10, 14, 19):
        try:
            for c in sh[row_idx]:
                c.fill = light_gray
                c.font = bold
        except Exception:
            pass
    sh["A1"] = "Order Summary"; sh["A1"].font = Font(bold=True, size=20)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in sh.iter_rows():
        for c in r: c.alignment = center

def _excel_bytes_ht(rows: List[Dict], restaurant_name: str) -> bytes:
    wb = Workbook()
    sh = wb.active
    _append_top_block(sh, restaurant_name)

    sales_cash = sales_bkash = sales_other = 0.0
    grand_subtotal = 0.0
    for r in rows:
        # Map straight from the HT Bitable fields
        order_date     = _ms_to_local_date_iso(r.get("Order Date"))
        order_id       = _norm_text(r.get("Order ID"))
        actual_item    = _num(r.get("Original Item Price"))
        item_price     = _num(r.get("Item Price"))
        discount       = _num(r.get("Discount"))
        special_ht     = _num(r.get("other discount by hungrytiger", 0))  # keep fallback
        bogo_pct       = _norm_text(r.get("BOGO Item Inflation Percentage")) or "0%"
        bogo_discount  = _num(r.get("BOGO Discount", 0))
        bogo_loss      = _num(r.get("BOGO Loss", 0))
        voucher_disc   = _num(r.get("Voucher discount", 0))
        voucher_code   = _norm_text(r.get("Voucher code") or "")
        rest_disc      = _num(r.get("Restaurant Discount"))
        ht_disc        = _num(r.get("Hungrytiger Discount"))
        bogo_rest_bear = _num(r.get("BOGO Discount bear by Restaurant", 0))
        bogo_ht_bear   = _num(r.get("BOGO Discount bear by Hungrytiger", 0))
        pay_type       = (_norm_text(r.get("Payment Type")) or "").upper()
        order_mode     = _norm_text(r.get("Order Mode"))
        tax            = _num(r.get("Tax"))
        sell_incl_tax  = _num(r.get("Selling price (inclusive of tax)"))
        cust_absorb    = _num(r.get("Customer absorb on delivery fees"))
        deliv_exp      = _num(r.get("Delivery fees expense"))
        comm_pct       = _norm_text(r.get("Commission Percentage") or "")
        comm_amt       = _num(r.get("Commission Amount"))
        svc_to_rest    = _num(r.get("Service fees to Restaurant"))
        svc_to_ht      = _num(r.get("Service fee to Hungrytiger"))
        tips_rest      = _num(r.get("Tips for restaurant"))
        bag_fees       = _num(r.get("Bag fees"))
        container_fees = _num(r.get("Container fees"))
        amt_to_rest    = _num(r.get("Amount to Restaurant"))
        ht_profit      = _num(r.get("HT Profit"))
        rest_name      = _norm_text(r.get("Restaurant") or restaurant_name)

        sh.append([
            order_date, order_id, f"{actual_item:.2f}", f"{item_price:.2f}", f"{discount:.2f}",
            f"{special_ht:.2f}", bogo_pct, f"{bogo_discount:.2f}", f"{bogo_loss:.2f}",
            f"{voucher_disc:.2f}", voucher_code, f"{rest_disc:.2f}", f"{ht_disc:.2f}",
            f"{bogo_rest_bear:.2f}", f"{bogo_ht_bear:.2f}", pay_type, order_mode,
            f"{tax:.2f}", f"{sell_incl_tax:.2f}", f"{cust_absorb:.2f}", f"{deliv_exp:.2f}",
            str(comm_pct), f"{comm_amt:.2f}", f"{svc_to_rest:.2f}", f"{svc_to_ht:.2f}",
            f"{tips_rest:.2f}", f"{bag_fees:.2f}", f"{container_fees:.2f}",
            f"{amt_to_rest:.2f}", f"{ht_profit:.2f}", rest_name
        ])

        grand_subtotal += amt_to_rest
        if pay_type == "CASH":
            sales_cash += amt_to_rest
        elif pay_type in {"BKASH", "BKASH_APP"}:
            sales_bkash += amt_to_rest
        else:
            sales_other += amt_to_rest

    # Footer summary
    sh.append([]); sh.append([])
    sh.append([""]*26 + [f"{grand_subtotal:.2f}"])
    sh.append([]); sh.append(["Payment"])
    sh.append(['', 'Comment', '', '', 'Sales through Cash on Delivery', '', '',
              'Sales Through bKash', '', 'Sales Through Other Method', '', '',
              '', 'Subtotal For Total Sales', '', '', 'Amount'])
    sh.append(['', 'Payments Deposited', '', '',
               f"{sales_cash:.2f}", '', '',
               f"{sales_bkash:.2f}", '',
               f"{sales_other:.2f}", '', '', '',
               f"{grand_subtotal:.2f}", '', '',
               f"{grand_subtotal:.2f}"])
    sh.append([]); sh.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "*Rounding Difference"])
    sh.append([]); sh.append(["Hungrytiger"])

    _style_sheet(sh)
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()

def generate_invoice_files_ht(
    rows: List[Dict], *, restaurant_name: str, start_ms: int, end_ms: int
) -> Tuple[str, str]:
    # Use nominal UTC+8 dates so labels match Control table exactly
    start_label = _nominal_label(start_ms)  # e.g. '2025-08-01'
    end_label   = _nominal_label(end_ms)    # e.g. '2025-08-12'

    rslug = _slug(restaurant_name)
    base_prefix = f"invoices/HT/{rslug}/{start_label}__{end_label}"
    x_name = f"{base_prefix}/HT_{rslug}_{start_label}__{end_label}.xlsx"

    excel_bytes = _excel_bytes_ht(rows, restaurant_name)  # keep as-is
    default_storage.save(x_name, ContentFile(excel_bytes))
    excel_url = default_storage.url(x_name)

    return excel_url, ""  # PDF handled elsewhere
